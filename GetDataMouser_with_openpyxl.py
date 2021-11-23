# -*- coding: utf-8 -*-
"""
Created on Wed Nov 11 10:27:28 2020

Получение данных о электронных компонентах с сайта TME через его API
Наименования (Артикулы) помещаются в первый столбец файла XLSX, без пропусков строк, можно с производителем через пробел.
"""
import openpyxl,time
from get_mouser_product_data import GetMouserProductData


class GetData:
    """Класс для получения информации по наименованию детали сайта TME"""
    # переменная-флаг для принудительного выхода из цикла поиска, используется графическим интерфейсом.
    flag=1
    
    @staticmethod
    # Функция считает все ячейки первого столбца в которых что то написано,до тех пор пока не встретит пустую ячейку "None"
    def number_of_articles(path):
        """Функция считает все ячейки первого столбца в которых что то написано, до тех пор пока не встретит пустую ячейку None"""
        # Открываем Эксель
        wb = openpyxl.load_workbook(path)#путь к файлу
        sheet = wb.active
        #выясняем количество артикулов в файле эксель
        i = 1
        while sheet['A'+str(i)].value != None:
            i+=1
        wb.save(path)
        return i-1
    @staticmethod
    # Функция создающая список артикулов. Принимает число артикулов и номер колонки в виде буквы (str): 'A'- первая колонка
    #
    def articlesList(n,column,path):
        """ Функция создающая список артикулов. Принимает число артикулов и номер колонки в виде буквы (str): 'A'- первая колонка"""
        # Открываем Эксель
        wb = openpyxl.load_workbook(path)#путь к файлу
        sheet = wb.active 
        cord_in = column+str(1)
        cord_out = column+str((n)) 
        # формирование списка артикулов
        vals = [v[0].value for v in sheet[cord_in : cord_out]] #[r[0].value for r in sheet.Range(cord)]
        return vals
        wb.save(path)
    
    def __init__(self, xlsxpath, API_KEY ):
        self.path = xlsxpath
        self.api_key = API_KEY
        
        self.GMPD = GetMouserProductData(self.api_key)

        self.quantity_of_articles = GetData.number_of_articles(self.path)
        self.articles_list = GetData.articlesList(self.quantity_of_articles,"A",self.path)


    # piece_of_data = {'APIdata': {'Errors': [], 'SearchResults': 
    #                               {'NumberOfResult': 3, 'Parts': [{'Availability': '5831 Склад в США', 'DataSheetUrl': '', 
    #                                                               'Description': 'LDO регуляторы напряжения 250-mA LDO', 
    #                                                               'FactoryStock': '0', 'ImagePath': 'https://www.mouser.com/images/texasinstruments/images/ITP_TI_SOIC-8_D_t.jpg',
    #                                                               'Category': 'LDO регуляторы напряжения', 'LeadTime': '70 Дни', 'LifecycleStatus': '', 'Manufacturer': 'Texas Instruments',
    #                                                               'ManufacturerPartNumber': 'TPS76618DR', 'Min': '1', 'Mult': '1', 'MouserPartNumber': '595-TPS76618DR',
    #                                                               'ProductAttributes': [{'AttributeName': 'Упаковка', 'AttributeValue': 'Cut Tape'},
    #                                                                                     {'AttributeName': 'Упаковка', 'AttributeValue': 'MouseReel'},
    #                                                                                     {'AttributeName': 'Упаковка', 'AttributeValue': 'Reel'},
    #                                                                                     {'AttributeName': 'Кол-во в стандартной упаковке',
    #                                                                                       'AttributeValue': '2500'}], 'PriceBreaks': [{'Quantity': 1, 'Price': '$2.24', 'Currency': 'USD'},
    #                                                                                                                                   {'Quantity': 10, 'Price': '$2.05', 'Currency': 'USD'},
    #                                                                                                                                   {'Quantity': 100, 'Price': '$1.48', 'Currency': 'USD'},
    #                                                                                                                                   {'Quantity': 250, 'Price': '$1.31', 'Currency': 'USD'}],
    #                                                                                                                                   'AlternatePackagings': [{'APMfrPN': 'TPS76618D'},
    #                                                                                                                                                           {'APMfrPN': ' TPS76618DG4'},
    #                                                                                                                                                           {'APMfrPN': ' TPS76618DRG4'}],
    #                                                                                                                                   'ProductDetailUrl': 'https://ru.mouser.com/ProductDetail/Texas-Instruments/TPS76618DR?qs=6zVL%252ByCp0mqVydlY%2FTLktg%3D%3D',
    #                                                                                                                                   'Reeling': True, 'ROHSStatus': 'RoHS Compliant', 'SuggestedReplacement': '', 'MultiSimBlue': 0,
    #                                                                                                                                   'UnitWeightKg': {'UnitWeight': 7.59e-05}, 'InfoMessages': [], 'ProductCompliance':
    #                                                                                                                                       [{'ComplianceName': 'CAHTS', 'ComplianceValue': '8542390029'},
    #                                                                                                                                       {'ComplianceName': 'CNHTS', 'ComplianceValue': '8542319000'},
    #                                                                                                                                       {'ComplianceName': 'USHTS', 'ComplianceValue': '8542390001'},
    #                                                                                                                                       {'ComplianceName': 'JPHTS', 'ComplianceValue': '8542390990'},
    #                                                                                                                                       {'ComplianceName': 'MXHTS', 'ComplianceValue': '85423999'},
    #                                                                                                                                       {'ComplianceName': 'TARIC', 'ComplianceValue': '8542399000'},
    #                                                                                                                                       {'ComplianceName': 'ECCN', 'ComplianceValue': 'EAR99'}]}]}},
    #                                                                                                                                   'ParserData': {'Производитель': 'Texas Instruments', 'Категория продукта': 'LDO регуляторы напряжения', 'RoHS': 'Подробности', 'Вид монтажа': 'SMD/SMT', 'Упаковка / блок': 'SOIC-8', 'Выходное напряжение': '1.8 V', 'Выходной ток': '250 mA', 'Количество выходов': '1 Output', 'Полярность': 'Positive', 'Ток покоя': '50 uA', 'Входное напряжение (макс.)': '10 V', 'Входное напряжение МИН.': '2.7 V', 'Тип выхода': 'Fixed', 'Минимальная рабочая температура': '- 40 C', 'Максимальная рабочая температура': '+ 125 C', 'Напряжение отпускания': '140 mV', 'Серия': 'TPS76618', 'Упаковка': 'Reel', 'Высота': '1.58 mm', 'Длина': '4.9 mm', 'Диапазон рабочих температур': '- 4', 'Ширина': '3.91 mm', 'Торговая марка': 'Texas Instruments', 'Точность регулирования напряжения': '3 %', 'Ib - Входной ток смещения': '35 uA', 'Нестабильность выходного напряжения или тока': '0.01 %/V', 'Нестабильность выходной нагрузки': '0.5 %', 'Рабочий ток источника питания': '30 uA', 'Pd - рассеивание мощности': '0.904 W', 'Тип продукта': 'LDO Voltage Regulators', 'Размер фабричной упаковки': '2500', 'Подкатегория': 'PMIC - Power Management ICs', 'Вес изделия': '76 mg', 'datasheetURL': 'http://www.ti.com/general/docs/suppproductinfo.tsp?distId=26&gotoUrl=http%3A%2F%2Fwww.ti.com%2Flit%2Fgpn%2Ftps766'}}


    def get_dict(self,part_name):
        """Функция использеет экшены, формирует и отдает словарь для одной детали. Получает на входе артикул детали."""
        line_data = {"Symbol": "","Description": None,"Weight": None,"Photo": "","ProductInformationPage": None,"OriginalSymbol": "","Producer":None, "ParameterList": "", "DocumentUrl": None}
        piece_of_data = self.GMPD.all_data(part_name)
        
        if piece_of_data == None:
            print("Кажется детали нет на MOUSER")
            return line_data
        
        elif piece_of_data != None and GetData.flag == 1:           
            print("Поиск детали")
            try:
                line_data["Symbol"] = piece_of_data['APIdata']['SearchResults']['Parts'][0]['MouserPartNumber']
            except IndexError:
                print(part_name, ": Название детали не найдено")
            try:
                line_data["Description"] = piece_of_data['APIdata']['SearchResults']['Parts'][0]['Description']
            except IndexError:
                print(part_name, ": Нет дескрипшина")
            try:
                photo = piece_of_data['APIdata']['SearchResults']['Parts'][0]['Description']
                
            except IndexError:
                print(part_name, ": Нет фото")    
            try:    
                line_data["ProductInformationPage"] = piece_of_data['APIdata']['SearchResults']['Parts'][0]['ProductDetailUrl']
            except IndexError:
                print(part_name, ": Нет ссылки на деталь")    
            try:   
                line_data["OriginalSymbol"] = piece_of_data['APIdata']['SearchResults']['Parts'][0]['ManufacturerPartNumber']
            except IndexError:
                print(part_name, ": Нет оригинального названия")   
            try:   
                line_data["Producer"] = piece_of_data['APIdata']['SearchResults']['Parts'][0]['Manufacturer']
            except IndexError:
                print(part_name, ": Нет производителя")      
            try:   
                line_data["Weight"] = piece_of_data['APIdata']['SearchResults']['Parts'][0]['UnitWeightKg']['UnitWeight']
            except KeyError:
                print(part_name,": Нет веса" )
            except IndexError:
                print(part_name,": Нет веса" )
            
            try:
                line_data["DocumentUrl"] = piece_of_data['ParserData']['datasheetURL']
            except TypeError:
                print('Не удалось получить даташит из массива парсера')
                try:
                    line_data["DocumentUrl"] = piece_of_data['APIdata']['SearchResults']['Parts'][0]['DataSheetUrl']
                except IndexError:
                    print("Нет даташита")
            
            try:
                
                parameters = piece_of_data['ParserData']
                print(parameters)
                # Удаляем ненужные поля
                parameters.pop('Производитель','Нет ключа производитель')
                parameters.pop('RoHS','Нет ключа')
                parameters.pop('Упаковка','Нет ключа')
                parameters.pop('Вес изделия','Нет ключа')
                parameters.pop('Торговая марка','Нет ключа')
                parameters.pop('datasheetURL','Нет ключа')
                parameters.pop('Подкатегория','Нет ключа')
                parameters.pop('Другие названия товара №','Нет ключа')
                parameters.pop('Тип продукта','Нет ключа')
                parameters.pop('Коммерческое обозначение','Нет ключа')
                parameters.pop('Ограничения на доставку','Нет ключа')
                parameters.pop('Размер фабричной упаковки','Нет ключа')
                parameters.pop('Чувствительный к влажности','Нет ключа')
                parameters.pop('Размер фабричной упаковкиРазмер фабричной упаковки','Нет ключа')
                parameters.pop('Тип ссылки','Нет ключа')
                #parameters.pop('Тип','Нет ключа')
                # Преобразуем словарь в текст
                #
                # strparam=[]
                # for i in list( parameters.items()):
                #     i=i[0]+' '+i[1].lower()
                #     strparam.append(i)
                # comma = ', '
                # paramstext = comma.join(strparam)
                #
                line_data["ParameterList"] = parameters
            except AttributeError:
                print(part_name,": Нет массива данных от парсера" )
            try:
                line_data["Photo"] = piece_of_data['APIdata']['SearchResults']['Parts'][0]['ImagePath']
            except IndexError:
                print('Нет ссылки на фото')
                line_data["Photo"] = None
            forced_wait=2
            print('Принудительное ожидание ', forced_wait, 'секунд:\n')
            for sec in range(forced_wait):
                print( sec + 1, 'сек.')
                time.sleep(1)   # Затормозим процесс для стабильности
                
            return line_data
            
            
    def duty_cycle(self, rng1=0):
        """ОСНОВНОЙ ЦИКЛ. Цикл проходящий по всем деталям из файла XLSX, заполняющая все графы."""
        self.quantity_of_articles = GetData.number_of_articles(self.path)
        self.articles_list = GetData.articlesList(self.quantity_of_articles,"A",self.path)
        # Открываем Эксель
        wb = openpyxl.load_workbook(self.path)#путь к файлу
        sheet = wb.active
        rng2=len(self.articles_list)
        index = 0
        for j in range(rng1,rng2):
            #global GetData.flag
            if GetData.flag != 1:
                break
            else:
                index += 1
                print('#',index)
                if self.articles_list[j] != None:
                    line_data = self.get_dict(self.articles_list[j])
                    sheet['B'+str(j+1)] = str(line_data["Symbol"])
                    sheet['C'+str(j+1)] = line_data["Description"]
                    sheet['D'+str(j+1)] = line_data["Weight"]
                    sheet['E'+str(j+1)] = line_data["Photo"]
                    sheet['F'+str(j+1)] = line_data["ProductInformationPage"]
                    sheet['G'+str(j+1)] = str(line_data["ParameterList"])
                    sheet['H'+str(j+1)] = line_data["DocumentUrl"]
                    sheet['K'+str(j+1)] = line_data["Producer"]
                    sheet['L'+str(j+1)] = str(line_data["OriginalSymbol"])
            wb.save(self.path)
        print("Цикл обращений к TME завершен.")
                                          
# Испытания      
if __name__ == '__main__':
    xlsxpath = "productdata.xlsx"
    API_KEY = "4aaaca99-d258-45ba-ad1d-aca303dd0a8a"

    A = GetData(xlsxpath,API_KEY)
    A.duty_cycle()
    #d = A.get_dict('TPS76618DR')
    #print(d)
